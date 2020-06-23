import openpyxl
import random


def getVal(pos):

    ref = [(1, 20), (2, 1), (3, 3), (4, 1), (5, 5), (6, 1), (7, 3), (8, 1), (9, 10), (10, 1), (11, 3), (12, 1), (13, 5), (14, 1), (15, 5), (16, 3), (17, 1), (18, 10), (19, 1), (20, 3), (21, 1), (22, 5), (23, 1), (24, 3), (25, 1)]

    for i in ref:

        if i[0] == pos:

            return i[1]

def readData():

    wb = openpyxl.load_workbook('gonzorust.xlsx')
    sheet = wb.active

    data = []

    column0 = []
    column1 = []
    column2 = []

    for i in sheet['A']:

        column0.append(i.value)

    for i in sheet['B']:

        column1.append(i.value)

    for x in sheet['C']:

        column2.append(x.value)

    for index, val in enumerate(column1):

        data.append((int(val), column2[index]))

    return column0

    

try:
    wb = openpyxl.load_workbook('gonzorust.xlsx')
except:
    wb = openpyxl.Workbook()
    wb.save(filename='gonzorust.xlsx')
sheet1 = wb.active
try:
    count = readData()[-1] + 1
except:
    count = 1

runningCount = 0
thirds = 0 

while True:
     
    pos = input("enter position: ")

    if pos != 'q':

        val = getVal(int(pos))

        if val != None:

            if val == 1 or val == 3:

                if thirds == 2:

                    thirds = 0
                    runningCount += 1
                else:

                    thirds += 1

            else:

                runningCount -= 1

            sheet1.cell(row=count, column=1).value = count
            sheet1.cell(row=count, column=2).value = pos
            sheet1.cell(row=count, column=3).value = val

        wb.save(filename='gonzorust.xlsx')

        print('your running count is '+str(runningCount))

        count += 1

    else:

        quit()


input()